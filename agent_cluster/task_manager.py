"""
任务仓库管理 + 输入持久化
核心设计原则：
  1. 模糊传递：OCR/解析失败不阻塞，用占位符标记，content.md 永远生成
  2. 快照存根：source/ 中的原始文件永不修改
  3. AI Translate 边界：content.md 是 Guard 的唯一文本入口
"""

import json
import shutil
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Literal

from .config import config
from .llm_client import LLMClient

# 时区从配置读取
_TZ = config.timezone

# ------------------------------------------------------------------
# 数据结构
# ------------------------------------------------------------------

ConversionStatus = Literal["full", "partial", "failed"]
TaskType = Literal["xlsx", "image_batch", "raw_text", "mixed"]


@dataclass
class Task:
    """任务统一抽象"""

    id: str
    type: TaskType
    folder_path: Path
    content_md: str = ""
    conversion_status: ConversionStatus = "full"
    input_text: str = ""
    source_files: list[str] = field(default_factory=list)
    created_at: str = ""

    @property
    def has_content(self) -> bool:
        return bool(self.content_md.strip())

    def content_preview(self, max_len: int = 200) -> str:
        return self.content_md[:max_len] + ("..." if len(self.content_md) > max_len else "")


# ------------------------------------------------------------------
# TaskManager
# ------------------------------------------------------------------


class TaskManager:
    """任务扫描、持久化、OCR 转换"""

    def __init__(self, llm_client: LLMClient | None = None):
        self._task_dir = Path(config.task_dir)
        self._task_dir.mkdir(parents=True, exist_ok=True)
        self._llm = llm_client or LLMClient()

    # ------------------------------------------------------------------
    # 输入持久化
    # ------------------------------------------------------------------

    def save_input(self, user_input: str, source_files: list[str] | None = None) -> str:
        """
        持久化用户输入到 task/{task_id}/ 独立文件夹。
        返回 task_id。
        """
        task_id = _generate_task_id()
        folder = self._task_dir / task_id
        folder.mkdir(parents=True, exist_ok=True)

        # 源文件子目录（快照存根）
        source_dir = folder / "source"
        source_dir.mkdir(exist_ok=True)

        saved_files: list[str] = []
        if source_files:
            for src in source_files:
                src_path = Path(src)
                if src_path.exists():
                    dst = source_dir / src_path.name
                    shutil.copy2(src_path, dst)
                    saved_files.append(src_path.name)

        # input.json
        input_data = {
            "text": user_input,
            "timestamp": datetime.now(_TZ).isoformat(),
            "source": "cli",
            "original_files": saved_files,
        }
        (folder / "input.json").write_text(
            json.dumps(input_data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        return task_id

    # ------------------------------------------------------------------
    # OCR → Markdown（模糊传递原则）
    # ------------------------------------------------------------------

    def ocr_to_md(self, task_id: str) -> ConversionStatus:
        """
        将 task 文件夹中的源文件转换为 content.md。
        模糊传递：逐文件尝试，失败用 [OCR_FAILED: ...] 占位，永不阻塞。
        扫描范围：source/ 子目录 + 任务文件夹根目录（用户可能直接拖入）。
        返回 conversion_status。
        """
        folder = self._task_dir / task_id
        source_dir = folder / "source"
        source_dir.mkdir(exist_ok=True)

        # 收集文件：source/ + 根目录（去重）
        _SKIP_NAMES = {"source", "content.md", "input.json", "result.json"}
        files: list[Path] = []
        seen: set[str] = set()

        for f in sorted(source_dir.iterdir()):
            if f.is_file():
                files.append(f)
                seen.add(f.name)

        for f in sorted(folder.iterdir()):
            if f.is_file() and f.name not in _SKIP_NAMES and f.name not in seen:
                files.append(f)
                seen.add(f.name)

        if not files:
            (folder / "content.md").write_text("", encoding="utf-8")
            return "full"

        parts: list[str] = []
        ok_count = 0
        fail_count = 0

        for f in files:
            if f.suffix.lower() in _IMAGE_EXTS:
                result = self._image_to_md(f)
            elif f.suffix.lower() in _XLSX_EXTS:
                result = self._xlsx_to_md(f)
            elif f.suffix.lower() in _TEXT_EXTS:
                result = self._text_to_md(f)
            else:
                result = f"[UNSUPPORTED: {f.name}]"

            if result.startswith("[OCR_FAILED:") or result.startswith("[UNSUPPORTED:"):
                fail_count += 1
            else:
                ok_count += 1

            parts.append(result)

        content = "\n\n---\n\n".join(parts)
        (folder / "content.md").write_text(content, encoding="utf-8")

        if fail_count == 0:
            return "full"
        elif ok_count == 0:
            return "failed"
        else:
            return "partial"

    # ------------------------------------------------------------------
    # 扫描 & 加载
    # ------------------------------------------------------------------

    def scan(self) -> list[Task]:
        """扫描 task/ 目录，返回所有任务摘要（不含 content_md 全文）"""
        tasks: list[Task] = []
        if not self._task_dir.exists():
            return tasks

        for entry in sorted(self._task_dir.iterdir(), reverse=True):
            if not entry.is_dir() or not entry.name.startswith("task_"):
                continue
            task = self.load(entry.name)
            if task:
                # 扫描时只返回摘要，content_md 仅保留前 200 字符
                task.content_md = task.content_preview(200) if task.content_md else "(empty)"
                tasks.append(task)

        return tasks

    def load(self, task_id: str) -> Task | None:
        """加载完整 Task 对象（含 content.md 全文）"""
        folder = self._task_dir / task_id
        if not folder.exists():
            return None

        # 读取 input.json
        input_file = folder / "input.json"
        input_text = ""
        source_files: list[str] = []
        created_at = ""
        if input_file.exists():
            try:
                data = json.loads(input_file.read_text(encoding="utf-8"))
                input_text = data.get("text", "")
                source_files = data.get("original_files", [])
                created_at = data.get("timestamp", "")
            except (json.JSONDecodeError, OSError):
                pass

        # 读取 content.md
        content_file = folder / "content.md"
        content_md = ""
        if content_file.exists():
            content_md = content_file.read_text(encoding="utf-8")

        # 推断 conversion_status
        conversion_status = self._infer_status(content_md)

        # 推断 task type
        task_type = self._infer_type(source_files, input_text)

        return Task(
            id=task_id,
            type=task_type,
            folder_path=folder,
            content_md=content_md,
            conversion_status=conversion_status,
            input_text=input_text,
            source_files=source_files,
            created_at=created_at,
        )

    # ------------------------------------------------------------------
    # 内部转换方法
    # ------------------------------------------------------------------

    def _image_to_md(self, path: Path) -> str:
        """PP-OCRv6 异步 Job API: 提交→轮询→提取 prunedResult"""
        if not config.paddleocr_api_url or not config.paddleocr_api_key:
            return f"[OCR_SKIP: {path.name} (PaddleOCR not configured)]"

        try:
            import time as _time
            import requests as _requests

            headers = {"Authorization": f"bearer {config.paddleocr_api_key}"}
            optional = json.dumps(
                {
                    "useDocOrientationClassify": False,
                    "useDocUnwarping": False,
                    "useTextlineOrientation": False,
                }
            )

            # Step 1: 提交 Job（multipart 上传）
            with open(path, "rb") as f:
                resp = _requests.post(
                    config.paddleocr_api_url,
                    headers=headers,
                    data={"model": "PP-OCRv6", "optionalPayload": optional},
                    files={"file": f},
                    timeout=30,
                )
            if resp.status_code != 200:
                return f"[OCR_FAILED: {path.name} (submit {resp.status_code}: {resp.text[:100]})]"

            job_id = resp.json()["data"]["jobId"]

            # Step 2: 轮询等待完成
            jsonl_url = ""
            for _ in range(60):
                _time.sleep(5)
                r = _requests.get(
                    f"{config.paddleocr_api_url}/{job_id}",
                    headers=headers,
                    timeout=10,
                )
                if r.status_code != 200:
                    continue
                state = r.json()["data"]["state"]
                if state == "done":
                    jsonl_url = r.json()["data"]["resultUrl"]["jsonUrl"]
                    break
                elif state == "failed":
                    msg = r.json()["data"].get("errorMsg", "unknown")
                    return f"[OCR_FAILED: {path.name} (job failed: {msg})]"
            else:
                return f"[OCR_FAILED: {path.name} (timeout)]"

            # Step 3: 下载 JSONL → 提取 prunedResult
            r = _requests.get(jsonl_url, timeout=30)
            r.raise_for_status()
            text = _extract_ocr_text_jsonl(r.text)
            if not text.strip():
                return f"[OCR_EMPTY: {path.name}]"
            return f"## {path.name}\n\n{text}"

        except Exception as e:
            return f"[OCR_FAILED: {path.name} ({e})]"

    def _xlsx_to_md(self, path: Path) -> str:
        """openpyxl 提取 Excel 文本→markdown 表格"""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, data_only=True)
            parts = [f"## {path.name}"]
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                parts.append(f"\n### Sheet: {sheet_name}\n")
                # 转为 markdown 表格
                max_cols = max(len(row) for row in rows)
                md_rows: list[str] = []
                for i, row in enumerate(rows):
                    cells = [str(c) if c is not None else "" for c in row]
                    cells += [""] * (max_cols - len(cells))
                    md_rows.append("| " + " | ".join(cells) + " |")
                    if i == 0:
                        md_rows.append("| " + " | ".join(["---"] * max_cols) + " |")
                parts.append("\n".join(md_rows))
            wb.close()
            return "\n\n".join(parts)
        except Exception:
            return f"[OCR_FAILED: {path.name}]"

    def _text_to_md(self, path: Path) -> str:
        """直接读取文本文件"""
        try:
            content = path.read_text(encoding="utf-8")
            return f"## {path.name}\n\n```\n{content}\n```"
        except Exception:
            return f"[OCR_FAILED: {path.name}]"

    # ------------------------------------------------------------------
    # 内部辅助
    # ------------------------------------------------------------------

    @staticmethod
    def _infer_status(content_md: str) -> ConversionStatus:
        if not content_md.strip():
            return "failed"
        if "[OCR_FAILED:" in content_md:
            # 全是占位符 = failed，有成功也有失败 = partial
            lines = content_md.strip().split("\n")
            fail_lines = sum(1 for l in lines if "[OCR_FAILED:" in l)
            content_lines = sum(1 for l in lines if l.startswith("##") and "[OCR_FAILED:" not in l)
            if content_lines == 0:
                return "failed"
            return "partial"
        return "full"

    @staticmethod
    def _infer_type(source_files: list[str], _input_text: str) -> TaskType:
        ext_set = {Path(f).suffix.lower() for f in source_files}
        has_img = bool(ext_set & _IMAGE_EXTS)
        has_xlsx = bool(ext_set & _XLSX_EXTS)
        if has_img and has_xlsx:
            return "mixed"
        if has_img:
            return "image_batch"
        if has_xlsx:
            return "xlsx"
        return "raw_text"


# ------------------------------------------------------------------
# 常量
# ------------------------------------------------------------------

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
_XLSX_EXTS = {".xlsx", ".xls"}
_TEXT_EXTS = {".txt", ".md", ".json", ".csv", ".yaml", ".yml", ".py", ".log"}


def _extract_ocr_text_jsonl(jsonl_text: str) -> str:
    """从 PP-OCRv6 JSONL 结果中提取纯文本。
    JSONL 每行: { "result": { "ocrResults": [ { "prunedResult": "识别文字" }, ... ] } }
    兼容 prunedResult 为字符串或嵌套 dict 的情况。
    """
    lines: list[str] = []
    for line in jsonl_text.strip().split("\n"):
        if not line.strip():
            continue
        try:
            data = json.loads(line)
            for res in data.get("result", {}).get("ocrResults", []):
                raw = res.get("prunedResult", "")
                if isinstance(raw, str) and raw.strip():
                    lines.append(raw.strip())
                elif isinstance(raw, dict):
                    # PP-OCRv6 可能返回 {"text": "...", ...} 嵌套结构
                    inner = raw.get("text", "") or raw.get("prunedResult", "")
                    if isinstance(inner, str) and inner.strip():
                        lines.append(inner.strip())
        except (json.JSONDecodeError, KeyError, AttributeError):
            pass
    return "\n\n".join(lines)


def _extract_ocr_text(data: dict) -> str:
    """旧版兼容兜底（非 JSONL 格式的 API）"""
    for key in ("result", "data"):
        items = data.get(key)
        if isinstance(items, list):
            lines = []
            for item in items:
                if isinstance(item, dict) and "text" in item:
                    lines.append(item["text"])
            if lines:
                return "\n".join(lines)
    if isinstance(data.get("text"), str) and data["text"].strip():
        return data["text"]
    return str(data)


def _generate_task_id() -> str:
    """生成 task_{YYYYMMDDTHHMMSS} 格式的 ID"""
    return f"task_{datetime.now(_TZ).strftime('%Y%m%dT%H%M%S')}"
